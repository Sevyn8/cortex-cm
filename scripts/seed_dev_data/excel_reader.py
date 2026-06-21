"""Excel reader: turn the seed workbook into structured Python data.

For each sheet in ``SHEET_MAPPINGS``, returns a list of dicts keyed
by column header. Helper-only columns (those starting with underscore:
``_key``, ``_tenant_key``, ``_parent_key``, ``_role_key``,
``_legal_name_FYI``, etc.) are kept in the dict; the loaders consume
them to resolve cross-sheet FK references via the UUIDMapper.

Two cell-classification helpers, intentionally asymmetric:

  ``_is_null_ish(v)``: STRICT. Used for per-cell translation during
  dict construction. True only when the cell is genuinely the
  data-author's "no value" signal: ``None``, empty/whitespace-only
  string, or a case variant of the literal string ``NULL`` (the
  seed Excel's README convention). Stays strict so that real rows
  with a mistaken ``#VALUE!`` cell preserve the error string in the
  resulting dict and surface as a loud INSERT failure at the
  loader. Hiding such cells silently would risk inserting rows that
  *look* well-formed but are missing data the author intended.

  ``_is_phantom_cell(v)``: BROAD. Used only for the row-skip check.
  True for everything ``_is_null_ish`` is true for, PLUS Excel
  formula-error sentinels (``#VALUE!``, ``#REF!``, ``#N/A``,
  ``#DIV/0!``, ``#NAME?``, ``#NULL!``, ``#NUM!``). This catches
  phantom rows that openpyxl's ``read_only=True`` mode sometimes
  yields at the worksheet's far-bottom (residual formulas in cells
  past the visible data, evaluating to ``#VALUE!``); such rows are
  not data the author wrote and are safely dropped.

The asymmetry matters. A phantom row entirely composed of error
sentinels is silently skipped; a real row whose author mis-typed a
single cell as ``#VALUE!`` keeps that string verbatim, the loader
attempts the INSERT, and Postgres rejects it loudly with the row's
context. Same effect at the data-quality level: errors surface
against real rows, never against phantoms.

Do not "unify" the two helpers without thinking through this trade.
"""
from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook  # type: ignore[import-untyped]

from scripts.seed_dev_data.column_mappings import SHEET_MAPPINGS


# All loadable sheets. Derived from SHEET_MAPPINGS to keep a single
# source of truth — when Step 6.2 lands audit_logs, adding it to
# SHEET_MAPPINGS automatically extends excel_reader's coverage too.
SHEETS_TO_READ: set[str] = set(SHEET_MAPPINGS.keys())

# Excel formula-error sentinels. A cell rendering one of these is a
# formula that couldn't compute. Treated as phantom for row-skip
# purposes only; preserved verbatim by per-cell translation so a
# mistaken error in a real row surfaces loudly at INSERT time.
_EXCEL_ERROR_SENTINELS: frozenset[str] = frozenset({
    "#VALUE!",
    "#REF!",
    "#N/A",
    "#DIV/0!",
    "#NAME?",
    "#NULL!",
    "#NUM!",
})


def _is_null_ish(value: Any) -> bool:
    """STRICT null-ness check. Used for per-cell translation.

    True for: ``None``; empty/whitespace-only string; case variants
    of the literal string ``NULL``. Nothing else. See module docstring
    for the asymmetry with ``_is_phantom_cell``.
    """
    if value is None:
        return True
    if isinstance(value, str):
        stripped = value.strip()
        if stripped == "" or stripped.upper() == "NULL":
            return True
    return False


def _is_phantom_cell(value: Any) -> bool:
    """BROAD phantom-cell check. Used for the row-skip filter only.

    True for everything ``_is_null_ish`` is true for, PLUS Excel
    formula-error sentinels (``#VALUE!``, etc.). See module docstring
    for the asymmetry with ``_is_null_ish``.
    """
    if _is_null_ish(value):
        return True
    if isinstance(value, str) and value.strip() in _EXCEL_ERROR_SENTINELS:
        return True
    return False


def read_workbook(path: Path) -> dict[str, list[dict[str, Any]]]:
    """Read all loadable sheets.

    Returns ``{sheet_name: [{col: value, ...}, ...]}``.
    """
    wb = load_workbook(path, read_only=True, data_only=True)
    result: dict[str, list[dict[str, Any]]] = {}
    for sheet_name in SHEETS_TO_READ:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            result[sheet_name] = []
            continue
        headers = list(rows[0])
        data_rows: list[dict[str, Any]] = []
        for row in rows[1:]:
            # Phantom-row skip: openpyxl's read_only mode sometimes
            # yields trailing rows past the visible data containing
            # residual formula errors (#VALUE!, etc.). The broader
            # _is_phantom_cell catches these; per-cell translation
            # below uses the strict _is_null_ish so any error
            # sentinel inside a real row surfaces loudly at INSERT.
            if all(_is_phantom_cell(v) for v in row):
                continue
            row_dict: dict[str, Any] = {}
            for h, v in zip(headers, row):
                if h is None:
                    continue
                row_dict[h] = None if _is_null_ish(v) else v
            data_rows.append(row_dict)
        result[sheet_name] = data_rows
    return result
